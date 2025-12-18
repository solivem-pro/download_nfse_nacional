import json
import os
import logging
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import Dict, Any
from datetime import datetime, timedelta
## Módulos auxiliares
from config.config import DIRETORIOS, ROOT_DIR
from config.utils import limpar_numero, formatar_milhar, formatar_cnpj_digitacao, validar_cnpj, limpar_cnpj, formatar_cnpj
from config.json_handler import carregar_json, salvar_json
from ui.ui_basic import modal_window, back_window, ToolTip, scrolled_treeview, buttons_frame, centralizar
logger = logging.getLogger(__name__)

## ------------------------------------------------------------------------------
## Interface principal de cadastros
## ------------------------------------------------------------------------------
class CadastroUI:
    """Interface para gerenciamento de cadastros de empresas"""
    
    def __init__(self, parent):
        self.parent = parent       
        self.win = None
        self.data = None
        self.chaves_cadastros = []
        self.tree = None
        self._setup_ui()

    def _setup_ui(self):
        """Configura a interface principal"""
        self.win = modal_window(self.parent.root, "Cadastros - Download NFSe Nacional", 800, 500)
        self.data = carregar_json(DIRETORIOS['cadastros_json'])
        
        self._criar_treeview()
        self._criar_botoes()
        self._atualizar_lista()


    def _criar_treeview(self):
        """Cria a treeview usando função do ui_basic"""
        columns_config = [
            ('cod', 'Código', 80, 'center', 'int'),  
            ('empresa', 'Empresa', 200, 'center', 'string'), 
            ('cnpj', 'CNPJ', 150, 'center', 'string'),  
            ('venc', 'Venc. Cert.', 100, 'center', 'date_dd_mm_yyyy')  
        ]
        self.tree, scrollbar, _ = scrolled_treeview(self.win, columns_config)
        
        # Configurar tags para formatação condicional
        self.tree.tag_configure('vencendo', foreground='red', font=('TkDefaultFont', 9, 'bold'))
        self.tree.tag_configure('normal', foreground='black')
        
        self.tree.bind('<<TreeviewSelect>>', self._on_select)

    def _criar_botoes(self):
        """Cria os botões da interface usando função do ui_basic"""
        botoes_config = [
            {'text': "Adicionar", 'width': 10, 'command': self._open_editor},
            {'text': "Editar", 'width': 10, 'state': tk.DISABLED, 'command': self._editar},
            {'text': "Editar NSU", 'width': 10, 'state': tk.DISABLED, 'command': self._editar_nsu},
            {'text': "Excluir", 'width': 10, 'state': tk.DISABLED, 'command': self._excluir},
            {'text': "Resetar NSUs", 'width': 12, 'command': self._resetar_nsu},
            {'text': "Excluir Todos", 'width': 15, 'command': self._excluir_todos}
        ]
        
        frame_buttons, botoes = buttons_frame(self.win, botoes_config)
        
        # Armazenar referências para atualização de estado
        self.btn_edit = botoes["Editar"]
        self.btn_nsu = botoes["Editar NSU"]
        self.btn_delete = botoes["Excluir"]
        
        # Botão voltar à direita
        btn_voltar = tk.Button(frame_buttons, text="Voltar", width=10, command=self._voltar)
        btn_voltar.pack(side=tk.RIGHT, padx=20)

    def _atualizar_lista(self):
        """Atualiza a lista de cadastros no treeview"""
        self.tree.delete(*self.tree.get_children())
        self.chaves_cadastros = []
        
        cadastros_ordenados = sorted(
            ((k, v) for k, v in self.data.items() if k.startswith("cadastro_") and k != "cadastro_0"),
            key=lambda item: item[1]["cod"]
        )
        
        for key, cadastro in cadastros_ordenados:
            vencimento = cadastro.get('venc', '')
            
            # Verificar se o certificado está próximo do vencimento
            tag = self._verificar_vencimento_proximo(vencimento)
            
            item = self.tree.insert('', tk.END, values=(
                cadastro['cod'],
                cadastro['empresa'],
                formatar_cnpj(cadastro['cnpj']),
                vencimento
            ), tags=(tag,))
            
            self.chaves_cadastros.append(key)
        
        self._atualizar_estado_botoes()

    def _verificar_vencimento_proximo(self, data_vencimento: str) -> str:
        """
        Verifica se a data de vencimento está dentro de 30 dias
        Retorna 'vencendo' se estiver próximo, 'normal' caso contrário
        """
        if not data_vencimento:
            return 'normal'
            
        try:
            # Converter string para datetime
            data_venc = datetime.strptime(data_vencimento, "%d/%m/%Y")
            data_hoje = datetime.now()
            
            # Calcular diferença em dias
            dias_para_vencer = (data_venc - data_hoje).days
            
            # Se estiver vencido ou a vencer em até 30 dias
            if dias_para_vencer <= 30:
                return 'vencendo'
                
        except ValueError:
            # Se houver erro na conversão da data, retorna normal
            pass
            
        return 'normal'

    def _atualizar_estado_botoes(self):
        """Atualiza o estado dos botões baseado na seleção"""
        has_selection = len(self.tree.selection()) > 0
        state = tk.NORMAL if has_selection else tk.DISABLED
        
        self.btn_edit.config(state=state)
        self.btn_nsu.config(state=state)
        self.btn_delete.config(state=state)

    def _on_select(self, event):
        """Manipula a seleção no treeview"""
        self._atualizar_estado_botoes()

    def _voltar(self):
        """Fecha a janela e retorna para a principal"""
        back_window(self.win, self.parent.root)

    def _open_editor(self, edit_key=None):
        """Abre o editor de cadastro"""
        EditorCadastro(self, edit_key)

    def _editar(self):
        """Abre o editor para editar cadastro existente"""
        sel = self.tree.selection()
        if not sel:
            return
            
        item = sel[0]
        index = self.tree.index(item)
        key = self.chaves_cadastros[index]
        self._open_editor(key)

    def _editar_nsu(self):
        """Abre o editor de NSU"""
        sel = self.tree.selection()
        if not sel:
            return
            
        item = sel[0]
        index = self.tree.index(item)
        key = self.chaves_cadastros[index]
        cadastro = self.data[key]
        
        EditorNSU(self, cadastro)

    def _resetar_nsu(self):
        """Reseta os NSUs de todas as empresas cadastradas"""
        if not messagebox.askyesno("Confirmação", "Deseja realmente resetar os NSUs de TODOS os cadastros?"):
            return

        logger.info(f"Limpando arquivos a partir de: {DIRETORIOS['packs']}")
            
        # Conteúdo a ser escrito nos arquivos
        conteudo_json = '''{
    "registros": {
        "2000": {
        "01": {"nsu_inicial": 0, "nsu_final": 0}
            }
        }
    }'''
            
        # Percorre recursivamente todas as pastas e subpastas
        for pasta_raiz, subpastas, arquivos in os.walk(DIRETORIOS['packs']):
            for arquivo in arquivos:
                if arquivo == "nsu_competencia.json":
                    caminho_arquivo = os.path.join(pasta_raiz, arquivo)
                    try:
                        # Abre o arquivo no modo de escrita e escreve o conteúdo JSON
                        with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                            f.write(conteudo_json)
                        logger.info(f"Conteúdo atualizado: {caminho_arquivo}")
                    except Exception as e:
                        logger.error(f"Erro ao atualizar {caminho_arquivo}: {str(e)}")

        salvar_json(self.data, DIRETORIOS['cadastros_json'])
        self._atualizar_lista()
        messagebox.showinfo("Sucesso", "NSUs de todos os cadastros foram resetados com sucesso!")

    def _excluir(self):
        """Exclui cadastro selecionado"""
        sel = self.tree.selection()
        if not sel:
            return

        item = sel[0]
        index = self.tree.index(item)
        key = self.chaves_cadastros[index]
        empresa_data = self.data[key]

        if not messagebox.askyesno("Confirmação", f"Deseja realmente excluir {empresa_data['empresa']}?"):
            return

        self._excluir_cadastro_completo(key, empresa_data)
        self._atualizar_lista()

    def _excluir_todos(self):
        """Exclui todos os cadastros"""
        if not messagebox.askyesno("Confirmação", "Deseja realmente excluir TODOS os cadastros?"):
            return

        keys = [k for k in self.data.keys() if k.startswith("cadastro_") and k != "cadastro_0"]
        for key in keys:
            self._excluir_cadastro_completo(key, self.data[key])

        self.data["cadastros"] = 1
        salvar_json(self.data, DIRETORIOS['cadastros_json'])
        self._atualizar_lista()
        messagebox.showinfo("Sucesso", "Todos os cadastros foram excluídos com sucesso!")

    def _excluir_cadastro_completo(self, key: str, empresa_data: Dict[str, Any]):
        """Exclui todos os arquivos e dados do cadastro"""
        cod_empresa = empresa_data['cod']

        # Deletar certificado
        cert_file = os.path.basename(empresa_data['cert_path'])
        if cert_file:
            cert_path = os.path.join(DIRETORIOS['certificados'], cert_file)
            if os.path.exists(cert_path):
                os.remove(cert_path)

        # Deletar pasta de notas
        notas_path = os.path.join(DIRETORIOS['notas'], str(cod_empresa))
        if os.path.exists(notas_path):
            shutil.rmtree(notas_path)

        # Deletar arquivo .zip
        zip_path = os.path.join(DIRETORIOS['notas'], f"{cod_empresa}.zip")
        if os.path.exists(zip_path):
            os.remove(zip_path)

        # Remover do JSON
        self.data.pop(key, None)
        salvar_json(self.data, DIRETORIOS['cadastros_json'])

## ------------------------------------------------------------------------------
## Interface do editor de cadastros
## ------------------------------------------------------------------------------
class EditorCadastro:
    """Editor para adicionar/editar cadastros de empresas"""
    logger = logging.getLogger(__name__)
    
    def __init__(self, parent, edit_key=None):
        self.parent = parent
        self.edit_key = edit_key
        self.fields = {}
        self.sub = None
        
        self._setup_ui()

    def _setup_ui(self):
        """Configura a interface do editor"""
        titulo = "Editar Cadastro" if self.edit_key else "Novo Cadastro"
        self.sub = modal_window(self.parent.win, f"{titulo} - Download NFSe Nacional", 460, 250)
        
        self._criar_campos()
        self._criar_botoes()
        
        if self.edit_key:
            self._preencher_dados()

    def _criar_campos(self):
        """Cria os campos do formulário"""
        labels = {
            "cod": "Código",
            "empresa": "Empresa", 
            "cnpj": "CNPJ",
            "cert_pass": "Senha Certificado",
            "cert_path": "Certificado (.pfx)",
            "venc": "Vencimento Cert."
        }

        tooltips = {
            "cod": "Código numérico único da empresa. Não pode ser alterado após criado.",
            "empresa": "Nome fantasia ou razão social da empresa.",
            "cnpj": "CNPJ da empresa (formatação automática).",
            "cert_pass": "Senha do certificado digital correspondente ao arquivo .PFX.",
            "cert_path": "Selecione o arquivo .PFX do certificado digital da empresa que será importado.",
            "venc": "Data de vencimento do certificado (será lida automaticamente ao importar o certificado)."
        }

        for i, (key, label) in enumerate(labels.items()):
            self.fields[key] = tk.StringVar()
            
            tk.Label(self.sub, text=label).grid(row=i, column=0, sticky="w", padx=8, pady=5)
            
            if key == "venc":
                # Campo de vencimento como readonly
                entry = tk.Entry(self.sub, textvariable=self.fields[key], width=45, state="readonly")
            else:
                entry = tk.Entry(self.sub, textvariable=self.fields[key], width=45)
                
            entry.grid(row=i, column=1, padx=8, pady=5)

            self._configurar_campo(key, entry, i)
            self._adicionar_tooltip(i, key, tooltips.get(key, ""))

    def _configurar_campo(self, key: str, entry: tk.Entry, row: int):
        """Configura comportamentos específicos para cada campo"""
        if key == "cod":
            entry.bind("<FocusOut>", self._validar_codigo)
            if self.edit_key:
                entry.config(state="readonly")
                
        elif key == "cnpj":
            entry.bind("<KeyRelease>", formatar_cnpj_digitacao)
            entry.bind("<FocusOut>", self._validar_cnpj)
            
        elif key == "cert_path":
            self._adicionar_botao_arquivo(entry, row)

        entry.bind("<Return>", lambda e: e.widget.tk_focusNext().focus())

    def _adicionar_botao_arquivo(self, entry, row: int):
        """Adiciona botão para selecionar arquivo"""
        btn_browse = tk.Button(self.sub, text="...", command=self._browse_file)
        btn_browse.grid(row=row, column=2, padx=3)

    def _adicionar_tooltip(self, row: int, key: str, text: str):
        """Adiciona tooltip ao campo"""
        if not text:
            return
            
        icon = tk.Label(self.sub, text="ℹ", fg="blue", cursor="question_arrow")
        icon.grid(row=row, column=3, sticky="w", padx=2)
        ToolTip(icon, text)

    def _criar_botoes(self):
        """Cria os botões de ação usando grid para manter consistência"""
        frame_actions = tk.Frame(self.sub)
        frame_actions.grid(row=6, column=0, columnspan=4, pady=15, sticky="ew")
        
        frame_actions.columnconfigure(0, weight=1)
        frame_actions.columnconfigure(1, weight=0)
        frame_actions.columnconfigure(2, weight=0)
        frame_actions.columnconfigure(3, weight=1)
        
        btn_salvar = tk.Button(frame_actions, text="Salvar", width=12, command=self._salvar)
        btn_salvar.grid(row=0, column=1, padx=5)
        
        btn_cancelar = tk.Button(frame_actions, text="Cancelar", width=12, 
                                 command=lambda: back_window(self.sub, self.parent.win))
        btn_cancelar.grid(row=0, column=2, padx=5)

    def _preencher_dados(self):
        """Preenche os campos com dados existentes para edição"""
        cadastro = self.parent.data[self.edit_key]
        for key in self.fields:
            # Converter código para string apenas para exibição
            value = cadastro.get(key, "")
            if key == "cod":
                value = str(value)  # Apenas para exibição no campo
            self.fields[key].set(value)

    def _browse_file(self):
        """Seleciona arquivo de certificado e lê data de vencimento"""
        path = filedialog.askopenfilename(parent=self.sub, filetypes=[("Certificados PFX", "*.pfx")])
        if not path:
            return

        cod_empresa = self.fields["cod"].get()
        password = self.fields["cert_pass"].get()
        if not cod_empresa:
            messagebox.showwarning("Código necessário", "Informe o código da empresa antes de importar o certificado.")
            return
        if not password:
            messagebox.showwarning("Senha necessária", "Informe a senha do certificado antes de importar o certificado.")
            return       

        destino_abs = os.path.join(DIRETORIOS['certificados'], f"{cod_empresa}.pfx")
        destino_rel = os.path.relpath(destino_abs, ROOT_DIR)

        try:
            shutil.copy2(path, destino_abs)
            self.fields["cert_path"].set(destino_rel)
            
            # Tentar ler a data de vencimento do certificado
            data_vencimento = self._ler_data_vencimento_certificado(destino_abs)
            if data_vencimento:
                self.fields["venc"].set(data_vencimento)
                messagebox.showinfo("Certificado", f"Certificado importado com sucesso.\nData de vencimento: {data_vencimento}")
            else:
                messagebox.showwarning("Certificado", "Certificado importado com sucesso.\nNão foi possível ler a data de vencimento.\nPossivelmente a senha está incorreta.")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao copiar certificado:\n{e}")

    def _ler_data_vencimento_certificado(self, cert_path: str) -> str:
        """
        Tenta ler a data de vencimento do certificado digital
        Retorna string vazia se não conseguir ler
        """
        try:
            # Tentativa 1: Usando cryptography (mais comum)
            try:
                from cryptography import x509
                from cryptography.hazmat.backends import default_backend
                from cryptography.hazmat.primitives.serialization import pkcs12
                
                senha = self.fields["cert_pass"].get().encode() if self.fields["cert_pass"].get() else None
                
                with open(cert_path, "rb") as f:
                    p12_data = f.read()
                
                # Carregar o certificado
                private_key, certificate, additional_certificates = pkcs12.load_key_and_certificates(
                    p12_data, senha, default_backend()
                )
                
                if certificate:
                    # Usar not_valid_after_utc para evitar o warning e converter para timezone local
                    data_venc_utc = certificate.not_valid_after_utc
                    # Converter para datetime sem timezone (naive) no fuso local
                    data_venc_local = data_venc_utc.astimezone().replace(tzinfo=None)
                    # Converter para formato brasileiro
                    return data_venc_local.strftime("%d/%m/%Y")
                    
            except ImportError:
                pass  # cryptography não disponível
            
            # Tentativa 2: Usando OpenSSL via subprocess
            try:
                import subprocess
                senha = self.fields["cert_pass"].get()
                
                # Comando para extrair informações do certificado
                cmd = [
                    'openssl', 'pkcs12', '-in', cert_path, 
                    '-clcerts', '-nokeys', '-passin', f'pass:{senha}'
                ]
                
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    # Extrair data de expiração do output
                    for line in result.stdout.split('\n'):
                        if 'notAfter=' in line:
                            # Formato: notAfter=Dec 31 23:59:59 2025 GMT
                            date_str = line.split('=')[1].strip()
                            # Converter para datetime
                            data_obj = datetime.strptime(date_str, '%b %d %H:%M:%S %Y GMT')
                            return data_obj.strftime("%d/%m/%Y")
                            
            except (ImportError, subprocess.TimeoutExpired, subprocess.SubprocessError, ValueError):
                pass  # OpenSSL não disponível ou comando falhou
            
            return ""  # Retorna vazio se nenhum método funcionou
            
        except Exception as e:
            self.logger.info(f"Erro ao ler data de vencimento do certificado: {e}")
            return ""

    def _validar_codigo(self, event):
        """Valida unicidade do código"""
        try:
            cod_digitado = int(self.fields["cod"].get())  # Já converte para int
        except ValueError:
            return

        for key, cadastro in self.parent.data.items():
            if not key.startswith("cadastro_") or key == self.edit_key:
                continue
            if cadastro["cod"] == cod_digitado:
                messagebox.showerror("Erro", f"O código {cod_digitado} já está em uso pela empresa {cadastro['empresa']}.")
                self.fields["cod"].set("")
                event.widget.focus_set()
                break

    def _validar_cnpj(self, event):
        """Valida unicidade do CNPJ"""
        cnpj_digitado = validar_cnpj(self.fields["cnpj"].get())
        if not cnpj_digitado:
            messagebox.showerror("Erro", "CNPJ inválido, deve conter 14 dígitos alfanuméricos.")
            event.widget.focus_set()
            return

        for key, cadastro in self.parent.data.items():
            if not key.startswith("cadastro_") or key == self.edit_key:
                continue
            if limpar_cnpj(cadastro["cnpj"]) == cnpj_digitado:
                messagebox.showerror("Erro", f"O CNPJ {formatar_cnpj(cnpj_digitado)} já está em uso pela empresa {cadastro['empresa']}.")
                self.fields["cnpj"].set("")
                event.widget.focus_set()
                break

    def _salvar(self):
        """Salva o cadastro"""
        if not self._validar_dados():
            return

        if self.edit_key:
            self._atualizar_cadastro()
        else:
            self._criar_cadastro()

        salvar_json(self.parent.data, DIRETORIOS['cadastros_json'])
        self.parent._atualizar_lista()
        back_window(self.sub, self.parent.win)

    def _validar_dados(self) -> bool:
        """Valida os dados do formulário"""
        empresa = self.fields["empresa"].get().strip()
        cod = self.fields["cod"].get()
        cnpj = validar_cnpj(self.fields["cnpj"].get())

        if not empresa:
            messagebox.showerror("Erro", "Informe o nome da empresa.")
            return False
        if not cod:
            messagebox.showerror("Erro", "Informe o código da empresa.")
            return False
        try:
            int(cod)  # Verifica se é um número válido
        except ValueError:
            messagebox.showerror("Erro", "Código deve ser um número inteiro.")
            return False
        if not cnpj:
            messagebox.showerror("Erro", "CNPJ inválido, deve conter 14 dígitos alfanuméricos.")
            return False

        return True

    def _atualizar_cadastro(self):
        """Atualiza cadastro existente"""
        for key, var in self.fields.items():
            # Converter código para inteiro ao salvar
            if key == "cod":
                self.parent.data[self.edit_key][key] = int(var.get())
            else:
                self.parent.data[self.edit_key][key] = var.get()
        self.parent.data[self.edit_key]["cnpj"] = limpar_cnpj(self.fields["cnpj"].get())
        messagebox.showinfo("Sucesso", "Cadastro atualizado com sucesso!")

    def _criar_cadastro(self):
        """Cria novo cadastro"""
        total = int(self.parent.data.get("cadastros", 1))
        new_key = f"cadastro_{total}"
        
        modelo = self.parent.data["cadastro_0"].copy()
        modelo.update({k: v.get() for k, v in self.fields.items()})
        # Converter código para inteiro
        modelo["cod"] = int(self.fields["cod"].get())
        modelo["cnpj"] = limpar_cnpj(self.fields["cnpj"].get())
        
        self.parent.data[new_key] = modelo
        self.parent.data["cadastros"] = total + 1

        self._criar_pasta_empresa(str(modelo["cod"]), modelo["cnpj"])  # Manter como string para caminhos de arquivo
        messagebox.showinfo("Sucesso", f"Cadastro da empresa criado com sucesso!")

    def _criar_pasta_empresa(self, cod_empresa: str, cnpj: str):
        """Cria pasta da empresa com estrutura inicial"""
        pasta_modelo = os.path.join(DIRETORIOS['notas'], '0')
        pasta_nova_empresa = os.path.join(DIRETORIOS['notas'], str(cod_empresa))
        
        if not os.path.exists(pasta_modelo):
            messagebox.showerror("Erro crítico", "Pasta modelo não encontrada. Contate o suporte.")
            return

        try:
            shutil.copytree(pasta_modelo, pasta_nova_empresa)
            self._modificar_arquivo_xlsm(pasta_nova_empresa, cnpj, cod_empresa)
        except Exception as e:
            messagebox.showwarning("Aviso", f"Erro ao copiar pasta modelo: {e}")

    def _modificar_arquivo_xlsm(self, pasta_empresa: str, cnpj: str, cod_empresa: str):
        """Modifica o arquivo .xlsm da empresa"""
        try:
            xlsm_files = [f for f in os.listdir(pasta_empresa) if f.endswith('.xlsm')]
            if not xlsm_files:
                messagebox.showwarning("Aviso", "Nenhum arquivo .xlsm encontrado na pasta modelo.")
                return
                
            xlsm_path = os.path.join(pasta_empresa, xlsm_files[0])
            novo_nome = os.path.join(pasta_empresa, f"relatorio_{cod_empresa}.xlsm")
            os.rename(xlsm_path, novo_nome)
            
            from openpyxl import load_workbook
            
            wb = load_workbook(novo_nome, keep_vba=True)
            if 'alvo' in wb.sheetnames:
                sheet = wb['alvo']
                sheet['A1'] = str(cnpj)
                sheet['A1'].number_format = '@'
                wb.save(novo_nome)
                
        except ImportError:
            messagebox.showwarning("Aviso", "Biblioteca openpyxl não disponível para modificar o arquivo .xlsm")
        except Exception as e:
            messagebox.showwarning("Aviso", f"Erro ao modificar arquivo .xlsm: {e}")

## ------------------------------------------------------------------------------
## Editor de NSU por competência
## ------------------------------------------------------------------------------
class EditorNSU:
    """Editor para gerenciar NSUs por competência"""

    def __init__(self, parent, cadastro: Dict[str, Any]):
        self.parent = parent
        self.cadastro = cadastro
        self.dados_nsu = {}
        self.win_nsu = None
        self.tree_nsu = None
        self.entry_ano = None
        self.entry_mes = None
        self.entry_inicial = None
        self.entry_final = None
        self.btn_delete_nsu = None
        self._setup_ui()

    def _setup_ui(self):
        """Configura a interface do editor de NSU"""
        self.win_nsu = modal_window(self.parent.win, 
                                     f"Editar NSU - {self.cadastro['empresa']} - Download NFSe Nacional", 
                                     500, 500)
        
        self._carregar_dados()
        self._criar_interface()

    def _carregar_dados(self):
        """Carrega os dados do arquivo NSU"""
        cod_empresa = self.cadastro['cod']
        pasta_empresa = os.path.join(DIRETORIOS['notas'], str(cod_empresa))
        arquivo_nsu = os.path.join(pasta_empresa, 'nsu_competencia.json')
        
        if not os.path.exists(arquivo_nsu):
            os.makedirs(pasta_empresa, exist_ok=True)
            self.dados_nsu = {
                "registros": {
                }
            }
            with open(arquivo_nsu, 'w', encoding='utf-8') as f:
                json.dump(self.dados_nsu, f, indent=2)
        else:
            with open(arquivo_nsu, 'r', encoding='utf-8') as f:
                self.dados_nsu = json.load(f)

    def _criar_interface(self):
        """Cria a interface do editor de NSU"""
        # Frame principal
        frame_main = tk.Frame(self.win_nsu)
        frame_main.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self._criar_treeview(frame_main)
        self._criar_controles()
        self._criar_botoes()
        self._atualizar_treeview()

    def _criar_treeview(self, parent):
        """Cria a treeview usando função do ui_basic"""
        columns_config = [
            ('ano', 'Ano', 80, 'center', 'int'),  
            ('mes', 'Mês', 80, 'center', 'int'),  
            ('nsu_inicial', 'NSU Inicial', 100, 'center', 'int'),  
            ('nsu_final', 'NSU Final', 100, 'center', 'int'),          ]
        self.tree_nsu, scrollbar, _ = scrolled_treeview(parent, columns_config)
        self.tree_nsu.bind('<<TreeviewSelect>>', self._on_tree_select)

    def _criar_controles(self):
        """Cria os campos de entrada e labels"""
        frame_controls = tk.Frame(self.win_nsu)
        frame_controls.pack(fill=tk.X, padx=10, pady=5)
        
        # Labels e campos
        tk.Label(frame_controls, text="Ano:").grid(row=0, column=0, padx=2, pady=2, sticky='w')
        self.entry_ano = tk.Entry(frame_controls, width=8)
        self.entry_ano.grid(row=0, column=1, padx=2, pady=2)
        
        tk.Label(frame_controls, text="Mês:").grid(row=0, column=2, padx=2, pady=2, sticky='w')
        self.entry_mes = tk.Entry(frame_controls, width=8)
        self.entry_mes.grid(row=0, column=3, padx=2, pady=2)
        
        tk.Label(frame_controls, text="NSU Inicial:").grid(row=0, column=4, padx=2, pady=2, sticky='w')
        self.entry_inicial = tk.Entry(frame_controls, width=8)
        self.entry_inicial.grid(row=0, column=5, padx=2, pady=2)
        
        tk.Label(frame_controls, text="NSU Final:").grid(row=0, column=6, padx=2, pady=2, sticky='w')
        self.entry_final = tk.Entry(frame_controls, width=8)
        self.entry_final.grid(row=0, column=7, padx=2, pady=2)
    
        
        # Configurar formatação numérica
        for entry in [self.entry_inicial, self.entry_final]:
            entry.bind("<KeyRelease>", formatar_milhar)
        
        # Bind dos campos para avançar com Enter
        campos = [self.entry_ano, self.entry_mes, self.entry_inicial, self.entry_final]
        for i, campo in enumerate(campos):
            proximo = campos[i+1] if i < len(campos)-1 else None
            if proximo:
                campo.bind("<KeyPress-Return>", lambda e, prox=proximo: self._avancar_campo(e, prox))
            else:
                campo.bind("<KeyPress-Return>", self._avancar_campo)

    def _criar_botoes(self):
        """Cria os botões de ação usando função do ui_basic"""
        botoes_config = [
            {'text': "Adicionar", 'width': 10, 'command': self._adicionar_registro},
            {'text': "Excluir", 'width': 10, 'command': self._excluir_registro, 'state': tk.DISABLED},
            {'text': "Excluir Todos", 'width': 15, 'command': self._excluir_todos_registros}
        ]
        
        frame_buttons, botoes = buttons_frame(self.win_nsu, botoes_config)
        
        self.btn_delete_nsu = botoes["Excluir"]
        
        # Botão voltar à direita
        btn_close = tk.Button(frame_buttons, text="Voltar", width=10, 
                            command=lambda: back_window(self.win_nsu, self.parent.win))
        btn_close.pack(side=tk.RIGHT, padx=20)

    def _atualizar_treeview(self):
        """Atualiza a treeview com os dados atuais"""
        self.tree_nsu.delete(*self.tree_nsu.get_children())
        registros = self.dados_nsu.get('registros', {})

        # Ordenar anos e meses
        anos_ordenados = sorted(registros.keys(), key=int, reverse=True)
        for ano in anos_ordenados:
            meses_ordenados = sorted(registros[ano].keys(), key=int, reverse=True)
            for mes in meses_ordenados:
                dados = registros[ano][mes]
                self.tree_nsu.insert('', tk.END, values=(
                    ano, 
                    str(mes).zfill(2),  # Força dois dígitos no mês
                    f"{dados.get('nsu_inicial', 0):,}".replace(",", "."),  # Formata com ponto de milhar
                    f"{dados.get('nsu_final', 0):,}".replace(",", "."),    # Formata com ponto de milhar
                ))

    def _limpar_campos(self):
        """Limpa os campos de entrada"""
        self.entry_ano.delete(0, tk.END)
        self.entry_mes.delete(0, tk.END)
        self.entry_inicial.delete(0, tk.END)
        self.entry_final.delete(0, tk.END)

    def _salvar_alteracoes(self):
        """Salva as alterações no arquivo JSON com formatação compacta"""
        cod_empresa = self.cadastro['cod']
        pasta_empresa = os.path.join(DIRETORIOS['notas'], str(cod_empresa))
        arquivo_nsu = os.path.join(pasta_empresa, 'nsu_competencia.json')
        
        try:
            with open(arquivo_nsu, 'w', encoding='utf-8') as f:
                # Formatação compacta - sem indentação
                json.dump(self.dados_nsu, f, ensure_ascii=False, separators=(',', ':'))
            messagebox.showinfo("Sucesso", "Alterações salvas com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar arquivo:\n{e}")

    def _adicionar_registro(self):
        """Adiciona novo registro de NSU"""
        ano = self.entry_ano.get().strip()
        mes = self.entry_mes.get().strip()
        nsu_inicial = limpar_numero(self.entry_inicial.get().strip())
        nsu_final = limpar_numero(self.entry_final.get().strip())
        
        # Validações
        if not ano or not mes:
            messagebox.showerror("Erro", "Ano e mês são obrigatórios!")
            return
        
        try:
            ano_int = int(ano)
            mes_int = int(mes)
            if mes_int < 1 or mes_int > 12:
                raise ValueError("Mês deve ser entre 1 e 12")
        except ValueError:
            messagebox.showerror("Erro", "Ano e mês devem ser números válidos!")
            return
        
        # Garantir estrutura - modificado para estrutura compacta
        if 'registros' not in self.dados_nsu:
            self.dados_nsu['registros'] = {}
        if ano not in self.dados_nsu['registros']:
            self.dados_nsu['registros'][ano] = {}
        
        # Adicionar/atualizar registro - versão compacta
        self.dados_nsu['registros'][ano][mes] = {
            "nsu_inicial": int(nsu_inicial) if nsu_inicial else 0,
            "nsu_final": int(nsu_final) if nsu_final else 0
        }
        
        self._salvar_alteracoes()
        self._limpar_campos()
        self._atualizar_treeview()

    def _excluir_registro(self):
        """Exclui registro selecionado"""
        selecionado = self.tree_nsu.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um registro para excluir!")
            return
        
        item = selecionado[0]
        valores = self.tree_nsu.item(item, 'values')
        
        if messagebox.askyesno("Confirmação", f"Excluir registro {valores[1]}/{valores[0]}?"):
            ano = valores[0]
            mes = valores[1]
            
            if ano in self.dados_nsu.get('registros', {}) and mes in self.dados_nsu['registros'][ano]:
                del self.dados_nsu['registros'][ano][mes]
                # Remover ano se ficar vazio
                if not self.dados_nsu['registros'][ano]:
                    del self.dados_nsu['registros'][ano]

            self._salvar_alteracoes()
            self._limpar_campos()
            self._atualizar_treeview()

    def _excluir_todos_registros(self):
        """Exclui todos os registros"""
        if messagebox.askyesno("Confirmação", "Deseja realmente excluir TODOS os registros?"):
            self.dados_nsu['registros'] = {}
            self._salvar_alteracoes()
            self._atualizar_treeview()

    def _on_tree_select(self, event):
        """Habilita/desabilita botões baseado na seleção"""
        selecionado = self.tree_nsu.selection()
        if selecionado:
            self.btn_delete_nsu.config(state=tk.NORMAL)
        else:
            self.btn_delete_nsu.config(state=tk.DISABLED)

    def _avancar_campo(self, event, proximo_campo=None):
        """Função para avançar para o próximo campo ou salvar se for o último"""
        if event.keysym == 'Return':
            # Verifica se todos os campos estão preenchidos
            campos = [self.entry_ano, self.entry_mes, self.entry_inicial, self.entry_final]
            todos_preenchidos = all(campo.get().strip() for campo in campos)
            
            if todos_preenchidos:
                # Se todos estão preenchidos, salva o registro
                self._adicionar_registro()
            elif proximo_campo:
                # Se há próximo campo, avança para ele
                proximo_campo.focus_set()