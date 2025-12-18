import os
import shutil
import time
from openpyxl import load_workbook
from config import DIRETORIOS

try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False
    print("AVISO: win32com nao esta instalado. A atualizacao do VBA sera ignorada.")
    print("Para habilitar atualizacao de VBA, instale: pip install pywin32")

BASE_DIR = os.path.join(DIRETORIOS["packs"], "0")
VBA_TXT_PATH = str(DIRETORIOS["vba"])  # Converter para string

def encontrar_xlsm_maior_nome(pasta):
    arquivos = [
        f for f in os.listdir(pasta)
        if f.lower().endswith(".xlsm")
    ]
    if not arquivos:
        return None
    return max(arquivos, key=lambda x: x.lower())

def ler_alvo_a1(caminho_xlsm):
    try:
        wb = load_workbook(caminho_xlsm, keep_vba=True, data_only=True)
        ws = wb["alvo"]
        valor = ws["A1"].value
        wb.close()
        return valor
    except Exception as e:
        print(f"  Erro ao ler alvo!A1 de {caminho_xlsm}: {e}")
        return None

def escrever_alvo_a1(caminho_xlsm, valor):
    try:
        wb = load_workbook(caminho_xlsm, keep_vba=True)
        ws = wb["alvo"]
        ws["A1"].value = valor
        wb.save(caminho_xlsm)
        wb.close()
        print(f"  Valor '{valor}' escrito em alvo!A1")
    except Exception as e:
        print(f"  Erro ao escrever em alvo!A1: {e}")

def atualizar_vba_base(caminho_xlsm):
    """Atualiza o VBA do arquivo base importando arquivo .bas"""
    if not WIN32COM_AVAILABLE:
        print("  win32com nao disponivel - pulando atualizacao de VBA")
        return False
    
    try:
        print(f"\n  Atualizando VBA do arquivo base...")
        
        # Converter VBA_TXT_PATH para string se necessario
        vba_path = str(VBA_TXT_PATH)
        
        # Verificar se arquivo .bas existe
        if not os.path.exists(vba_path):
            print(f"  Arquivo .bas nao encontrado: {vba_path}")
            return False
        
        # Verificar extensao .bas
        if not vba_path.lower().endswith('.bas'):
            print(f"  Arquivo VBA nao e .bas: {vba_path}")
            return False
        
        print(f"  Arquivo .bas encontrado: {os.path.basename(vba_path)}")
        
        # Iniciar Excel
        print("  Iniciando Excel...")
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Abrir arquivo
        print(f"  Abrindo arquivo base...")
        caminho_absoluto = os.path.abspath(caminho_xlsm)
        wb = excel.Workbooks.Open(caminho_absoluto)
        
        # Acessar projeto VBA (pode exigir habilitar confianca no acesso ao modelo de objeto VBA)
        print("  Acessando projeto VBA...")
        vb_proj = wb.VBProject
        
        # Constantes do VBA para tipos de componentes
        vbext_ct_StdModule = 1  # Modulo padrao
        vbext_ct_ClassModule = 2  # Classe
        vbext_ct_MSForm = 3  # Formulario
        vbext_ct_ActiveXDesigner = 11  # Designer ActiveX
        
        # Lista de componentes para remover
        componentes_para_remover = []
        
        print("  Removendo modulos VBA existentes...")
        
        # Primeiro, coletar todos os componentes que nao sao do tipo Document
        for i in range(vb_proj.VBComponents.Count, 0, -1):
            componente = vb_proj.VBComponents.Item(i)
            tipo_componente = componente.Type
            
            # Manter apenas os componentes do tipo Document (ThisWorkbook, Planilhas)
            # Remover todos os outros tipos
            if tipo_componente in [vbext_ct_StdModule, vbext_ct_ClassModule, vbext_ct_MSForm, vbext_ct_ActiveXDesigner]:
                print(f"    Removendo: {componente.Name} (tipo: {tipo_componente})")
                vb_proj.VBComponents.Remove(componente)
        
        print(f"  Todos os modulos VBA anteriores foram removidos")
        
        # Importar novo modulo .bas
        modulo_nome = os.path.basename(vba_path).replace('.bas', '')
        print(f"  Importando modulo: {modulo_nome}")
        
        # Verificar se o arquivo .bas existe antes de importar
        if os.path.exists(vba_path):
            vb_proj.VBComponents.Import(vba_path)
            print(f"  Modulo importado com sucesso")
        else:
            print(f"  ERRO: Arquivo .bas nao encontrado para importacao: {vba_path}")
            wb.Close()
            excel.Quit()
            return False
        
        print(f"  Salvando alteracoes de VBA...")
        wb.Save()
        wb.Close()
        
        # Fechar Excel
        excel.Quit()
        
        # Garantir que o processo do Excel seja finalizado
        time.sleep(1)
        
        print(f"  VBA atualizado com sucesso! Todos os modulos anteriores foram removidos.")
        return True
        
    except Exception as e:
        print(f"  Erro ao atualizar VBA: {e}")
        # Tentar fechar o Excel em caso de erro
        try:
            excel.Quit()
        except:
            pass
        return False

def obter_xlsm_base():
    print(f"\nProcurando arquivo base em: {BASE_DIR}")
    
    for f in os.listdir(BASE_DIR):
        if f.lower().endswith(".xlsm"):
            caminho = os.path.join(BASE_DIR, f)
            print(f"  Arquivo base encontrado: {f}")
            
            # Atualizar VBA do arquivo base
            atualizar_vba_base(caminho)
            
            return caminho
    
    raise FileNotFoundError("Arquivo .xlsm base nao encontrado na pasta 0")

def atualizar_pasta(pasta, xlsm_base):
    if pasta == BASE_DIR:
        print(f"  Pasta base - pulando")
        return

    antigo_nome = encontrar_xlsm_maior_nome(pasta)
    if not antigo_nome:
        print(f"  Nenhum arquivo .xlsm encontrado")
        return

    print(f"  Arquivo existente: {antigo_nome}")
    antigo_path = os.path.join(pasta, antigo_nome)
    alvo_valor = ler_alvo_a1(antigo_path)

    if alvo_valor is None:
        print(f"  Nao foi possivel ler o valor de alvo!A1")
        return

    print(f"  Valor de alvo!A1: '{alvo_valor}'")

    # Criar copia atualizada
    novo_temp = os.path.join(pasta, "__TEMP_UPDATE__.xlsm")
    print(f"  Criando copia atualizada...")
    shutil.copy2(xlsm_base, novo_temp)

    # Escrever valor antigo
    escrever_alvo_a1(novo_temp, alvo_valor)

    # Remover arquivo antigo
    print(f"  Removendo arquivo antigo...")
    os.remove(antigo_path)

    # Renomear para nome original
    novo_final = os.path.join(pasta, antigo_nome)
    os.rename(novo_temp, novo_final)
    
    print(f"  Pasta atualizada: {os.path.basename(pasta)}")

def atualizar_todas():
    print("=" * 60)
    print("INICIANDO ATUALIZACAO DE PLANILHAS")
    print("=" * 60)
    
    try:
        # Obter e atualizar arquivo base
        print(f"\n1. PREPARANDO ARQUIVO BASE")
        xlsm_base = obter_xlsm_base()
        
        print(f"\n2. REPLICANDO PARA OUTRAS PASTAS")
        
        # Listar apenas pastas imediatas dentro de /packs
        # Ignorar subpastas dentro de cada pasta (como /1515/TOMADOS, /1515/PRESTADOS, etc.)
        pastas_encontradas = []
        
        for item in os.listdir(DIRETORIOS["packs"]):
            caminho = os.path.join(DIRETORIOS["packs"], item)
            if os.path.isdir(caminho):
                pastas_encontradas.append(caminho)
        
        print(f"  Total de pastas encontradas no nivel raiz: {len(pastas_encontradas)}")
        
        # Atualizar cada pasta (exceto a base que ja foi processada)
        for i, caminho in enumerate(pastas_encontradas, 1):
            print(f"\n[{i}/{len(pastas_encontradas)}] Processando: {os.path.basename(caminho)}")
            atualizar_pasta(caminho, xlsm_base)
        
        print("\n" + "=" * 60)
        print("ATUALIZACAO CONCLUIDA COM SUCESSO!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\nERRO CRITICO: {e}")
        print("Processo interrompido.")

if __name__ == "__main__":
    print("Script de Atualizacao de Planilhas .xlsm")
    print(f"Diretorio base: {BASE_DIR}")
    print(f"Caminho VBA: {VBA_TXT_PATH}")
    
    if not os.path.exists(BASE_DIR):
        print(f"ERRO: Diretorio base nao encontrado: {BASE_DIR}")
        print("Pressione ENTER para fechar.")
        input()
        exit(1)
    
    # Verificar se o arquivo VBA existe
    if WIN32COM_AVAILABLE and not os.path.exists(VBA_TXT_PATH):
        print(f"AVISO: Arquivo VBA nao encontrado: {VBA_TXT_PATH}")
    
    atualizar_todas()
    
    print("\n" + "=" * 60)
    print("Script concluido. Mantendo terminal aberto...")
    print("\nPressione ENTER para fechar ou Ctrl+C para sair imediatamente.")
    
    try:
        # Manter script aberto
        input()
    except KeyboardInterrupt:
        print("\nScript encerrado pelo usuario.")